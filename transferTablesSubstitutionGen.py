#!/usr/bin/env python2
# -*- coding: utf-8 -*-

import pandas as pd
import openpyxl as xl
from modules import search_and_sort
from modules import format_dataframe
from modules import dfsearch_insert_cols
from modules import write_sub_file_first
from modules import write_sub_file_mid
from modules import write_sub_file_last

if hasattr(__builtins__, 'raw_input'):
    input=raw_input

# prompt user for file name of Workbook
fileName  =   input('Enter file name of XLSX File: ')
# prompt user for BLEPS beamline name
beamLine    =   '{' + input('Enter beamline name: ') + ':,'
substitutionFile = 'bleps.substitution.gen'
# open workbook
book = xl.load_workbook(fileName, data_only=True)

# create list of sheet names
sheetNames = book.sheetnames
# initialize dictionary for storing process variables keyed by sheet name
sheetDict = {}
# row index constant for assigning dictionary index according to process variable attribute (type, base name, pv name, epics ethernet tag, description)
cRo = 1
# initialize dictionary of DataFrames
dataframeDict = {}
indexList  = []

#
recordTypes = ['Int', 'Bool', 'Iint', 'Dint', 'Real']
# list of new column orders for PV data frames
newOrder = ['PV Name', 'EPICS Ethernet Tag', 'Short Description', 'Base Name', 'Type']
# for binary input: scan, zero name, one name, zero severity, one severity
biOrder  = ['{P,', 'N', 'TAG', 'SCAN', 'ZNAM', 'ONAM', 'ZSV', 'OSV', 'DESC}' ]
# for binary output: high, zero name, one name
boOrder  = ['{P', 'N', 'TAG', 'HIGH', 'ZNAM', 'ONAM', 'DESC}']
# for analog input: scan, precision, engineering units, hihi, hi, lo, lolo, hihi + hi severity, lolo + lo severity
aiOrder  = ['{P', 'N', 'TAG', 'SCAN', 'PREC', 'EGU', 'HIHI', 'HIGH', 'LOW', 'LOLO', 'HHSV', 'HSV', 'LSV', 'LLLSV', 'DESC}']
# arrange columns titles for each data type
columnsList = [aiOrder, biOrder, aiOrder, aiOrder, aiOrder]
dbFilename = ['bleps_ai.db', 'bleps_bi.db', 'bleps_ai.db', 'bleps_ai.db', 'bleps_ai.db']

# pull each entry from each sheet in workbook into an intermediary dictionary 'currDict', appending this current dictionary to the end of a list of all sheets/dictionary 'dictList' 
for s in range(len(sheetNames)): # iterate over each sheet of workbook
    dictList = [] # initialize/re-initialize list of the process variables used on the current sheets
    maxRow = book[sheetNames[s]].max_row # find the max height of each sheet; i.e. the last row in each sheet
    maxCol = 5     # column index constant (each process variable has 5 attributes listed previously)
    currRow = 2    # set current row index for each page iteration (start at row 2, as row 1 is column titles/process variable attribute title ; rows are 1-indexed )
    for row in book[sheetNames[s]].iter_rows(2,maxRow,1,1,True):     # iterate over column A for all process variables 
        currDict = {} # re/initalize current dictionary containing the current process variable's attributes
        currCol = 1   # set current column index for each row/process variable iteration ( start at column 1, 'Type' attribute ; columns are 0-indexed )
        for cell in row: # check each process variable's 'Used' field for 'X'
            if cell in ['x','X']: # if 'X' character found in 'Used' field, add row to current dictionary
                for col in book[sheetNames[s]].iter_cols(1,maxCol,currRow,currRow,True): # iterate over used process variable's attributes/columns
                    currDict[book[sheetNames[0]][cRo][currCol].value] = book[sheetNames[s]][currRow][currCol].value # add current attribute to current dictionary, using attribute name/column title as dictionary key
                    currCol += 1 # increment column index to add next attribute
                dictList.append(currDict) # add process variable's attribute-dictionary to list of all process variables
            currRow += 1 # increment row index to check if next process variable is used
    sheetDict[sheetNames[s]] = dictList # dictList now contains current sheet's used process variables ; add to dictionary of all used process variables, keyed by sheet name

i = 0
while i < len( sheetNames ): # remove dictionary keys/pages that have no process variables present
    if not sheetDict[ sheetNames[i] ]:
        del sheetDict[ sheetNames[i] ]
        del sheetNames[i]
    i += 1  
    
for i in range(len(sheetNames)): # convert each process variable dictionary/workbook sheet to a DataFrame, store in dictionary keyed by sheet name
    dataframeDict[sheetNames[i]] = pd.DataFrame(sheetDict[ sheetNames[i]])
    newColumns = newOrder + (dataframeDict[ sheetNames[i]].columns.drop(newOrder).tolist()) # create list of columns in the order required for the substitution file
    dataframeDict[sheetNames[i]] = dataframeDict[sheetNames[i]][newColumns] # replace columns with this new required column list
    dataframeDict[sheetNames[i]].insert(0,'{P,', str(beamLine)) # insert column holding beamline identifier
    dataframeDict[ sheetNames[i] ] = dataframeDict[ sheetNames[i] ].rename(columns={"PV Name":"N", "EPICS Ethernet Tag":"TAG", "Short Description":"DESC}"}) # rename existing columns to name required by substitution file 
    
# 'Display' sheet of transfer table workbook has non-homogeneous 'Type' field (i.e. PVs of several data types)
# therefore, sort the 'Display' dataframe alphabetically by 'Type' field
sortReturn = search_and_sort(dataframeDict['Display'],'Type') 
if sortReturn != 1: # function returns value '1' if Type field is homogeneous, therefore look for none-1 return value
    dataframeDict['Display'] = sortReturn[0] # save sorted dataframe back to dictionary
    changeIndex = sortReturn[1] # save list of tuples that point to last entry of each Type 

# write each workbook sheet dataframe to substitution file
for i in range(len(sheetNames)): # for all sheets    
    lastFlag = False
    firstFlag = False # set flags for first/last PV
    skipFlag = False
    typeFlag = 'none'
    for j in range(len(dataframeDict[sheetNames[i]])): # for all process variables in each sheet
        if j == len(dataframeDict[sheetNames[i]])-1:
            lastFlag = True
        if j == 0:
            firstFlag = True  # check if first or last process variable to be written
        currEntry = pd.DataFrame(dataframeDict[sheetNames[i]].loc[j,:]).transpose() # create intermediate dataframe from dataframe dictionary, transpose rows to columns (using 'loc' and pd.Dataframe transposes columns to rows)
        
        # check if current sheet dataframe is the 'Display' sheet
        if sheetNames[i] == 'Display' and sortReturn != 1: # process Display sheet first if Type field is non-homogeneous (expected case)
            skipFlag = True # if Display sheet is currently being processed set skipFlag to true so that Display data frame isn't processed again in subsequent code
            jj = 0  # manual index for something
            if currEntry.at[currEntry.index[0],'Type'] == 'Bool': # if current Type is associated with binary input records, use bi columns
                dfsearch_insert_cols(currEntry,currEntry.index.values[0],'Type','Bool',biOrder,insertLocation=3)
                currEntry = format_dataframe(currEntry)
                typeFlag = 'Bool'
            elif currEntry.at[currEntry.index[0],'Type'] == 'Int': # only other expected case is Type associated with analog input records, use ai columns
                dfsearch_insert_cols(currEntry,currEntry.index.values[0],'Type','Int',aiOrder,insertLocation=3)
                currEntry = format_dataframe(currEntry)
                typeFlag = 'Int'
            if firstFlag or currEntry.index[0] == changeIndex[0][0]+1: 
            # if first process variable of current sheet to be written to substitution file or the current entry just changed data type
            # write db file declaration at head of sheet section to substitution file
                if typeFlag == 'Bool': # if data type is of analog input type, write ai db invocation
                    write_sub_file_first(substitutionFile,sheetNames[i],'bleps_ai.db',currEntry)
                elif typeFlag == 'Int': # if not ai type, only other case is binary input type; write bi db invocation
                    write_sub_file_first(substitutionFile,sheetNames[i],'bleps_bi.db',currEntry)
                    firstFlag = False # after writing first entry of section, set firstFlag to false
                    break # go to next 
            elif lastFlag or currEntry.index[0] == changeIndex[0][0]: 
                # if last process variable of current sheet to be written to substitution file or the current entry is about to change data type
                write_sub_file_last(substitutionFile,currEntry) # write closing brackets for db invocation, write appropriate line breaks for next sheets section
                lastFlag = False # after writing last entry of section, set lastFlag to false
                break
            else:
                write_sub_file_mid(substitutionFile,currEntry) # if current PV is neither first nor last, must be a middle entry, write to sub file without special invocation or brackets
                break

        # check if current sheet dataframe is the 'EPICS_Inputs' sheet ;  this sheet is composed of boolean data types, but utilizes binary output data type                      
        if sheetNames[i] == 'EPICS_Inputs': 
            skipFlag = True # if EPICS_Inputs sheet is currently being processed set skipFlag to true so that EPICS_Inputs data frame isn't processed again in subsequent code
            dfsearch_insert_cols(currEntry,j,'Type','Bool',boOrder,insertLocation=3)
            currEntry = format_dataframe(currEntry)
            if firstFlag: # if first process variable of current sheet to be written to substitution file
                # write db file declaration at head of sheet section to substitution file
                write_sub_file_first(substitutionFile,sheetNames[i],"bleps_bo.db",currEntry)                
                firstFlag = False # after writing first entry of section, set firstFlag to false
                continue # go to next
            elif lastFlag:  # if last process variable of current sheet to be written to substitution file
                write_sub_file_last(substitutionFile,currEntry)                                    
                lastFlag = False # after writing last entry of section, set lastFlag to false
                continue # go to next
            else:
                write_sub_file_mid(substitutionFile,currEntry) # if current PV is neither first nor last, must be a middle entry, write to sub file without special invocation or brackets           
                continue
            continue  # move to next process variable
        
        if not skipFlag: # if current dataframe sheet is neither Display or EPICS_Inputs, proceed through subsequent code to process the rest of the workbook sheet dataframes  
            for ij in range( len(recordTypes) ): # iterate thru all possible record types
                # if current recordType matches the Type field of the entries of the current sheet dataframe, dfsearch_insert_cols returns TRUE
                if dfsearch_insert_cols(currEntry, j, 'Type', recordTypes[ij], columnsList[ij], insertLocation=3):    
                    currEntry = format_dataframe(currEntry) # on TRUE case, current dataframe has appropriate columns inserted, next function formats the entries to those columns with commas, brackets as needed
                    if firstFlag: # if first process variable to be written to substitution file
                        # write db file declaration at head of sheet section to substitution file
                        write_sub_file_first(substitutionFile,sheetNames[i],dbFilename[ij],currEntry)
                        firstFlag = False # after writing first entry of section, set firstFlag to false
                        break # go to next
                    elif lastFlag: # if last process variable
                        # write closing brackets for db invocation, write appropriate line breaks for next sheets section
                        write_sub_file_last(substitutionFile,currEntry)                     
                        lastFlag = False
                        break
                    else:
                        write_sub_file_mid(substitutionFile,currEntry)
                        break
                # if entry in the Type field does not match the list of expected Types (most likely because of transcription error), 
                # display to user the erroneous entry and prompt to select valid data type
                elif ij == (len(recordTypes)-1):
                    print "Erroneous data type in workbook\nData type for current process variable:", currEntry.at[j,'Type'] 
                    print "Choose record type from list:\n1.)", recordTypes[0], "\n2.)", recordTypes[1], "\n3.)", recordTypes[2], "\n4.)", recordTypes[3], "\n5.)", recordTypes[4]
                    while True:            
                        dt = int(input("Enter 1-5: "))
                        if dt in [1,2,3,4,5]:
                            dfsearch_insert_cols(currEntry,j,'Type',currEntry.at[j,'Type'],columnsList[dt-1],insertLocation=3)
                            currEntry = format_dataframe(currEntry)
                            if firstFlag: # if first process variable to be written to substitution file
                                write_sub_file_first(substitutionFile,sheetNames[i],dbFilename[ij],currEntry)
                                firstFlag = False
                                break # go to next
                            elif lastFlag: # if last process variable
                                write_sub_file_last(substitutionFile,currEntry)
                                lastFlag = False
                                break
                            else:
                                write_sub_file_mid(substitutionFile,currEntry)
                                break
                            break
                        else:
                            print "Wrong entry."
                            continue